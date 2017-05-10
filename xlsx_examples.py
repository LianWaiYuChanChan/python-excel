from openpyxl import load_workbook
from openpyxl import Workbook

def write():
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    # now we'll fill it with 100 rows x 200 columns

    for irow in range(100):
        ws.append(['%d' % i for i in range(200)])
    # save the file
    wb.save('zgenerated_new_big_file.xlsx')

def change_one_cell():
    wb = load_workbook(filename='sample.xlsx')
    ws = wb.get_sheet_by_name('Sheet1')
    ws.cell(row=4, column=8).value = 'WOrld'
    # Can save it to itself: sample.xlsx
    wb.save('zgenerated_sample_changed.xlsx')
    wb.close()

def read():
    wb = load_workbook(filename='sample.xlsx')
    sheet_ranges = wb['Sheet1']
    sheet_ranges['C5'] = "World"
    print(sheet_ranges['B5'].value)
    wb.close()


if __name__ == '__main__':
    read()
    write()
    change_one_cell()



